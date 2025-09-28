from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, time

app = Flask(__name__)
CORS(app)

FILE_PATH = "VEHICLE COUNTER.xlsx"

# ===============================
# Excel Template
# ===============================
def create_excel_template():
    """Create a clean Excel file without formatting issues"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Counter Data"
    
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_font = Font(bold=True)

    # Light vehicles header
    ws['A1'] = 'LIGHT VEHICLES'
    ws['A1'].font = bold_font
    ws['A1'].fill = header_fill

    ws['A2'] = 'Time'
    ws['A2'].font = bold_font
    ws['A2'].fill = header_fill

    approaches_light = [
        ('C2', 'North Approach', 'C2:F2'),
        ('G2', 'East Approach', 'G2:J2'), 
        ('K2', 'South Approach', 'K2:N2'),
        ('O2', 'West Approach', 'O2:R2')
    ]
    
    for cell, title, merge_range in approaches_light:
        ws[cell] = title
        ws[cell].font = bold_font
        ws[cell].fill = header_fill
        ws.merge_cells(merge_range)

    ws['A3'] = 'Period Start'
    ws['B3'] = 'Period End'
    ws['A3'].font = bold_font
    ws['A3'].fill = header_fill  
    ws['B3'].font = bold_font
    ws['B3'].fill = header_fill
    
    light_movements = ['U', 'R', 'SB', 'L', 'U', 'R', 'WB', 'L', 'U', 'R', 'NB', 'L', 'U', 'R', 'EB', 'L']
    light_cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
    
    for col, movement in zip(light_cols, light_movements):
        cell = f'{col}3'
        ws[cell] = movement
        ws[cell].font = bold_font
        ws[cell].fill = header_fill

    # Heavy vehicles header
    ws['T1'] = 'HEAVY VEHICLES'
    ws['T1'].font = bold_font
    ws['T1'].fill = header_fill
    
    ws['T2'] = 'Time'
    ws['T2'].font = bold_font
    ws['T2'].fill = header_fill

    approaches_heavy = [
        ('V2', 'North Approach', 'V2:Y2'),
        ('Z2', 'East Approach', 'Z2:AC2'),
        ('AD2', 'South Approach', 'AD2:AG2'), 
        ('AH2', 'West Approach', 'AH2:AK2')
    ]
    
    for cell, title, merge_range in approaches_heavy:
        ws[cell] = title
        ws[cell].font = bold_font
        ws[cell].fill = header_fill
        ws.merge_cells(merge_range)
    
    ws['T3'] = 'Period Start'
    ws['U3'] = 'Period End'
    ws['T3'].font = bold_font
    ws['T3'].fill = header_fill
    ws['U3'].font = bold_font  
    ws['U3'].fill = header_fill
    
    heavy_movements = ['U', 'R', 'SB', 'L', 'U', 'R', 'WB', 'L', 'U', 'R', 'NB', 'L', 'U', 'R', 'EB', 'L']
    heavy_cols = ['V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK']
    
    for col, movement in zip(heavy_cols, heavy_movements):
        cell = f'{col}3'
        ws[cell] = movement
        ws[cell].font = bold_font
        ws[cell].fill = header_fill
    
    all_light_cols = light_cols
    all_heavy_cols = heavy_cols
    
    for i in range(96):
        row = i + 4
        start_hour = (i * 15) // 60
        start_min = (i * 15) % 60
        end_hour = ((i + 1) * 15) // 60
        end_min = ((i + 1) * 15) % 60
        
        if end_hour >= 24:
            end_hour = 0
            
        start_time = f'{start_hour:02d}:{start_min:02d}'
        end_time = f'{end_hour:02d}:{end_min:02d}'
        
        ws[f'A{row}'] = start_time
        ws[f'B{row}'] = end_time
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        
        ws[f'T{row}'] = start_time
        ws[f'U{row}'] = end_time
        ws[f'T{row}'].fill = header_fill
        ws[f'U{row}'].fill = header_fill
        
        for col in all_light_cols:
            ws[f'{col}{row}'] = 0
            
        for col in all_heavy_cols:
            ws[f'{col}{row}'] = 0
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15) 
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(FILE_PATH)
    print(f"Clean Excel template created successfully at {FILE_PATH}")
    return wb

# ===============================
# Helper functions
# ===============================
def find_time_period_row(ws, period_text):
    if "From" not in period_text:
        return None
    try:
        parts = period_text.split("From ")[1].split(" to ")
        if len(parts) != 2:
            return None
        start_time = parts[0].strip()
        end_time = parts[1].strip()
        for row in range(4, 100):
            cell_start = ws[f'A{row}'].value
            cell_end = ws[f'B{row}'].value
            if cell_start and cell_end:
                if str(cell_start).strip() == start_time and str(cell_end).strip() == end_time:
                    return row
            if cell_start is None:
                break
        return None
    except:
        return None

def get_approach_columns(approach, is_heavy=False):
    if is_heavy:
        approach_map = {
            'North': ['V', 'W', 'X', 'Y'],
            'East': ['Z', 'AA', 'AB', 'AC'],
            'South': ['AD', 'AE', 'AF', 'AG'],
            'West': ['AH', 'AI', 'AJ', 'AK']
        }
    else:
        approach_map = {
            'North': ['C', 'D', 'E', 'F'],
            'East': ['G', 'H', 'I', 'J'],
            'South': ['K', 'L', 'M', 'N'],
            'West': ['O', 'P', 'Q', 'R']
        }
    return approach_map.get(approach, None)

# ===============================
# Routes
# ===============================
@app.route("/")
def home():
    return render_template("dual.html")

@app.route("/single")
def single_page():
    return render_template("single.html")

@app.route("/save", methods=["POST"])
def save_data():
    try:
        data = request.get_json()
        if not os.path.exists(FILE_PATH):
            create_excel_template()
        
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        
        period = data.get("period", "")
        approach = data.get("approach", "").replace(" Approach", "").replace("Not Set", "")
        
        if not period or not approach or period == "Not Set" or approach == "Not Set":
            return jsonify({"message": "Please set both time period and approach before saving!"}), 400
        
        row = find_time_period_row(ws, period)
        if not row:
            return jsonify({"message": f"Could not find time period '{period}' in Excel template!"}), 400
        
        light_columns = get_approach_columns(approach, False)
        heavy_columns = get_approach_columns(approach, True)
        
        if not light_columns or not heavy_columns:
            return jsonify({"message": f"Invalid approach '{approach}'!"}), 400
        
        light_data = data["light"]
        ws[f'{light_columns[0]}{row}'].value = light_data["uTurn"]
        ws[f'{light_columns[1]}{row}'].value = light_data["right"]
        ws[f'{light_columns[2]}{row}'].value = light_data["straight"]
        ws[f'{light_columns[3]}{row}'].value = light_data["left"]
        
        heavy_data = data["heavy"]
        ws[f'{heavy_columns[0]}{row}'].value = heavy_data["uTurn"]
        ws[f'{heavy_columns[1]}{row}'].value = heavy_data["right"]
        ws[f'{heavy_columns[2]}{row}'].value = heavy_data["straight"]
        ws[f'{heavy_columns[3]}{row}'].value = heavy_data["left"]
        
        wb.save(FILE_PATH)
        
        return jsonify({"message": f"Data saved successfully for {approach} approach, period {period}!"})
    except Exception as e:
        return jsonify({"message": f"Error saving data: {str(e)}"}), 500

@app.route("/test")
def test_connection():
    return jsonify({"message": "Flask server is running!"})

@app.route("/check-periods")
def check_periods():
    try:
        if not os.path.exists(FILE_PATH):
            return jsonify({"message": "Excel file doesn't exist. Create template first."})
        
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        periods = []
        for row in range(4, 20):
            start = ws[f'A{row}'].value
            end = ws[f'B{row}'].value
            if start and end:
                start_str = str(start).strip()
                end_str = str(end).strip()
                periods.append({"row": row, "start": start_str, "end": end_str, "period_string": f"From {start_str} to {end_str}"})
            else:
                break
        return jsonify({"periods": periods})
    except Exception as e:
        return jsonify({"message": f"Error checking periods: {str(e)}"}), 500

@app.route("/create-template", methods=["POST"])
def create_template_endpoint():
    try:
        if os.path.exists(FILE_PATH):
            os.remove(FILE_PATH)
        create_excel_template()
        return jsonify({"message": "Clean Excel template created successfully!"})
    except Exception as e:
        return jsonify({"message": f"Error creating template: {str(e)}"}), 500

@app.route("/debug-columns")
def debug_columns():
    try:
        result = {
            "light_vehicles": {
                "North": get_approach_columns("North", False),
                "East": get_approach_columns("East", False),
                "South": get_approach_columns("South", False),
                "West": get_approach_columns("West", False)
            },
            "heavy_vehicles": {
                "North": get_approach_columns("North", True),
                "East": get_approach_columns("East", True),
                "South": get_approach_columns("South", True),
                "West": get_approach_columns("West", True)
            }
        }
        return jsonify(result)
    except Exception as e:
        return jsonify({"message": f"Error: {str(e)}"}), 500

# ===============================
# Run
# ===============================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host="0.0.0.0", port=port)
